"""
report_generator.py
-------------------
Produces the final ``allocation_output.xlsx`` workbook:

  Sheet "Section <X>"   — one sheet per section (A, B, C, …)
  Sheet "Course Summary"  — enrolment vs capacity per course-section
  Sheet "Unallocated Students" — students not allocated + reason
"""

import logging
import re
from collections import defaultdict

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Colour palette ────────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_ALT_ROW_FILL = PatternFill("solid", fgColor="D6E4F0")
_WARN_FILL = PatternFill("solid", fgColor="FFF3CD")   # amber  – partial alloc
_ERROR_FILL = PatternFill("solid", fgColor="FFCDD2")  # red    – unallocated
_FULL_FILL = PatternFill("solid", fgColor="FFE0B2")   # orange – section full
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


class ReportGenerator:
    """Builds and saves the allocation Excel report."""

    def __init__(self, output_path: str = "allocation_output.xlsx"):
        self.output_path = output_path

    # ── Public API ────────────────────────────────────────────────────────────

    def generate_reports(
        self,
        allocated: list[dict],
        unallocated: list[dict],
        section_manager,
        allocation_mode: str = "auto",
    ) -> str:
        """
        Build the workbook and save it to ``self.output_path``.

        Returns the path that was written.
        """
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # discard the default blank sheet
        section_label_map: dict[tuple[str, str], str] = {}
        single_mode = allocation_mode == "single"

        combo_summary = self._build_combo_sections(
            allocated,
            section_manager.courses_config,
            single_mode,
        )

        # ── Single sorted allocated list ─────────────────────────────────────
        def _section_sort_value(student: dict) -> tuple[int, str]:
            label = self._section_label_for_student(student, section_label_map)
            try:
                return (0, f"{int(str(label)) :06d}")
            except Exception:
                return (1, str(label or ""))

        sorted_allocated = sorted(
            allocated,
            key=lambda s: (
                _section_sort_value(s),
                str(s.get("reg_no", "")),
            ),
        )
        self._write_allocated_sheet(wb, sorted_allocated, section_label_map, single_mode)

        # ── Course Summary ────────────────────────────────────────────────────
        self._write_course_summary(wb, combo_summary)

        # ── Unallocated ───────────────────────────────────────────────────────
        sorted_unallocated = sorted(unallocated, key=lambda s: str(s.get("reg_no", "")))
        self._write_unallocated_sheet(wb, sorted_unallocated)

        wb.save(self.output_path)
        logger.info("Report saved → %s", self.output_path)
        return self.output_path

    # ── Sheet writers ─────────────────────────────────────────────────────────

    def _write_section_sheet(
        self,
        wb: openpyxl.Workbook,
        sheet_name: str,
        students: list[dict],
        section_label_map: dict[tuple[str, str], str],
    ) -> None:
        ws = wb.create_sheet(sheet_name)
        headers = [
            "#",
            "Name",
            "Registration Number",
            "Email",
            "CGPA",
            "Allocated Primary Course (G1)",
            "Allocated Secondary Course (G2)",
            "Section Number",
        ]
        self._write_header_row(ws, headers)

        for idx, student in enumerate(students, start=1):
            row_data = [
                idx,
                student.get("name", ""),
                student.get("reg_no", ""),
                student.get("email", ""),
                student.get("cgpa", ""),
                student.get("allocated_primary") or "—",
                student.get("allocated_secondary") or "—",
                self._section_label_for_student(student, section_label_map),
            ]
            ws.append(row_data)
            row_num = ws.max_row

            # Determine row fill
            if not student.get("allocated_secondary"):
                fill = _WARN_FILL   # partial allocation
            elif idx % 2 == 0:
                fill = _ALT_ROW_FILL
            else:
                fill = None

            self._style_data_row(ws, row_num, len(headers), fill)

        self._set_column_widths(ws, [4, 24, 18, 30, 8, 22, 22, 14])
        ws.freeze_panes = "A2"
        logger.info("Sheet '%s' written — %d student(s).", sheet_name, len(students))

    def _write_course_summary(
        self,
        wb: openpyxl.Workbook,
        combo_summary: list[dict],
    ) -> None:
        ws = wb.create_sheet("Course Summary")
        headers = [
            "Primary Course (G1)",
            "Secondary Course (G2)",
            "Section Number",
            "Enrolled",
            "Capacity",
            "Available",
            "Fill %",
        ]
        self._write_header_row(ws, headers)

        for row in combo_summary:
            ws.append(
                [
                    row.get("g1", ""),
                    row.get("g2", ""),
                    row.get("section", ""),
                    row.get("enrolled", 0),
                    row.get("capacity", 0),
                    row.get("available", 0),
                    f"{row.get('fill_pct', 0)}%",
                ]
            )
            row_num = ws.max_row
            fill = _FULL_FILL if row.get("available", 0) == 0 else None
            self._style_data_row(ws, row_num, len(headers), fill)

        self._set_column_widths(ws, [28, 28, 14, 10, 10, 10, 10])
        ws.freeze_panes = "A2"

    def _write_unallocated_sheet(
        self, wb: openpyxl.Workbook, unallocated: list[dict]
    ) -> None:
        ws = wb.create_sheet("Unallocated Students")
        headers = ["Name", "Registration Number", "Email", "CGPA", "Reason"]
        self._write_header_row(ws, headers)

        for student in unallocated:
            ws.append(
                [
                    student.get("name", ""),
                    student.get("reg_no", ""),
                    student.get("email", ""),
                    student.get("cgpa", ""),
                    student.get("unallocated_reason", "Unknown"),
                ]
            )
            self._style_data_row(ws, ws.max_row, len(headers), _ERROR_FILL)

        self._set_column_widths(ws, [24, 18, 30, 8, 40])
        ws.freeze_panes = "A2"
        logger.info("Unallocated sheet written — %d student(s).", len(unallocated))

    def _write_allocated_sheet(
        self,
        wb: openpyxl.Workbook,
        allocated: list[dict],
        section_label_map: dict[tuple[str, str], str],
        single_mode: bool = False,
    ) -> None:
        ws = wb.create_sheet("Allocated Students")
        if single_mode:
            headers = [
                "Name",
                "Registration Number",
                "Email",
                "CGPA",
                "Allocated Course (G1)",
                "Section Number",
            ]
        else:
            headers = [
                "Name",
                "Registration Number",
                "Email",
                "CGPA",
                "Allocated Primary Course (G1)",
                "Allocated Secondary Course (G2)",
                "Section Number",
            ]
        self._write_header_row(ws, headers)

        for idx, student in enumerate(allocated, start=1):
            if single_mode:
                ws.append(
                    [
                        student.get("name", ""),
                        student.get("reg_no", ""),
                        student.get("email", ""),
                        student.get("cgpa", ""),
                        student.get("allocated_primary") or "—",
                        self._section_label_for_student(student, section_label_map),
                    ]
                )
            else:
                ws.append(
                    [
                        student.get("name", ""),
                        student.get("reg_no", ""),
                        student.get("email", ""),
                        student.get("cgpa", ""),
                        student.get("allocated_primary") or "—",
                        student.get("allocated_secondary") or "—",
                        self._section_label_for_student(student, section_label_map),
                    ]
                )
            fill = _ALT_ROW_FILL if idx % 2 == 0 else None
            self._style_data_row(ws, ws.max_row, len(headers), fill)

        if single_mode:
            self._set_column_widths(ws, [24, 18, 30, 8, 28, 14])
        else:
            self._set_column_widths(ws, [24, 18, 30, 8, 24, 24, 14])
        ws.freeze_panes = "A2"
        logger.info("Allocated sheet written — %d student(s).", len(allocated))

    def _build_section_label_map(self, section_manager) -> dict[tuple[str, str], str]:
        """
        Create plain integer labels 1, 2, 3 ... in alphabetical course +
        section order.
        """
        summary = section_manager.get_section_summary()
        label_map: dict[tuple[str, str], str] = {}
        counter = 1

        for course_name in sorted(summary.keys(), key=lambda x: x.lower()):
            sections = sorted(summary[course_name].keys())
            for section in sections:
                label_map[(course_name, section)] = str(counter)
                counter += 1

        return label_map

    def _section_label_for_student(
        self,
        student: dict,
        section_label_map: dict[tuple[str, str], str],
    ) -> str:
        section = student.get("allocated_section")
        if not section:
            return ""

        primary = student.get("allocated_primary")
        if primary and (primary, section) in section_label_map:
            return section_label_map[(primary, section)]

        secondary = student.get("allocated_secondary")
        if secondary and (secondary, section) in section_label_map:
            return section_label_map[(secondary, section)]

        # Final fallback to avoid blank section numbers in exported sheet.
        return str(section)

    def _build_combo_sections(
        self,
        allocated: list[dict],
        courses_config: dict[str, dict],
        single_mode: bool,
    ) -> list[dict]:
        """
        Build sections from combination groups with same-G1 compaction.

        Rules:
          - group by (G1, G2) in dual/auto mode and by (G1, '') in single mode
          - split each (G1, G2) group by combo capacity first
          - then compact chunks across different G2 for the same G1
            when combined size does not exceed G1 section capacity
          - assign section numbers after groups are formed
        """
        groups: dict[tuple[str, str], list[dict]] = defaultdict(list)
        for s in allocated:
            g1 = str(s.get("allocated_primary") or "").strip()
            g2 = "" if single_mode else str(s.get("allocated_secondary") or "").strip()
            if not g1:
                continue
            groups[(g1, g2)].append(s)

        combo_summary: list[dict] = []
        section_no = 1
        chunks_by_g1: dict[str, list[dict]] = defaultdict(list)

        for (g1, g2) in sorted(groups.keys(), key=lambda x: (x[0].lower(), x[1].lower())):
            students = sorted(groups[(g1, g2)], key=lambda s: str(s.get("reg_no", "")))

            g1_cap = int(courses_config.get(g1, {}).get("capacity", 60) or 60)
            if single_mode or not g2:
                combo_capacity = g1_cap
            else:
                g2_cap = int(courses_config.get(g2, {}).get("capacity", g1_cap) or g1_cap)
                combo_capacity = min(g1_cap, g2_cap)
            combo_capacity = max(1, combo_capacity)

            for i in range(0, len(students), combo_capacity):
                chunk = students[i:i + combo_capacity]
                if chunk:
                    chunks_by_g1[g1].append({
                        "g2": g2,
                        "students": chunk,
                    })

        for g1 in sorted(chunks_by_g1.keys(), key=lambda x: x.lower()):
            g1_cap = max(1, int(courses_config.get(g1, {}).get("capacity", 60) or 60))
            chunks = sorted(
                chunks_by_g1[g1],
                key=lambda c: (-len(c["students"]), str(c.get("g2", ""))),
            )

            bins: list[dict] = []

            # Best-fit placement: merge different G2 chunks under same G1
            # whenever the combined strength remains within G1 capacity.
            for chunk in chunks:
                size = len(chunk["students"])
                best_idx = None
                best_left = None

                for idx, b in enumerate(bins):
                    left = g1_cap - (len(b["students"]) + size)
                    if left < 0:
                        continue
                    if best_left is None or left < best_left:
                        best_left = left
                        best_idx = idx

                if best_idx is None:
                    bins.append({"students": [], "g2_list": []})
                    best_idx = len(bins) - 1

                bins[best_idx]["students"].extend(chunk["students"])
                g2 = str(chunk.get("g2", "")).strip()
                if g2:
                    bins[best_idx]["g2_list"].append(g2)

            for b in bins:
                label = str(section_no)
                students_in_bin = sorted(b["students"], key=lambda s: str(s.get("reg_no", "")))
                for student in students_in_bin:
                    student["allocated_section"] = label

                g2_values = sorted({x for x in b["g2_list"] if x})
                if single_mode:
                    g2_text = ""
                elif not g2_values:
                    g2_text = ""
                elif len(g2_values) == 1:
                    g2_text = g2_values[0]
                else:
                    g2_text = " + ".join(g2_values)

                enrolled = len(students_in_bin)
                available = g1_cap - enrolled
                fill_pct = round((enrolled / g1_cap) * 100, 1)

                combo_summary.append(
                    {
                        "g1": g1,
                        "g2": g2_text,
                        "section": label,
                        "enrolled": enrolled,
                        "capacity": g1_cap,
                        "available": available,
                        "fill_pct": fill_pct,
                    }
                )
                section_no += 1

        # ── Apply admin section overrides (set after combo labels are stable) ──
        has_admin = any(s.get("_admin_section") for s in allocated)
        if has_admin:
            for s in allocated:
                forced = s.pop("_admin_section", None)
                if forced:
                    s["allocated_section"] = str(forced)

            # Rebuild enrolled counts in the summary to reflect moves.
            from collections import Counter
            counts: Counter = Counter(
                (str(s.get("allocated_section", "")), str(s.get("allocated_primary", "")).strip())
                for s in allocated
            )
            for entry in combo_summary:
                new_enrolled = counts.get((str(entry["section"]), entry["g1"]), 0)
                entry["enrolled"] = new_enrolled
                cap = entry["capacity"]
                entry["available"] = cap - new_enrolled
                entry["fill_pct"] = round(new_enrolled / cap * 100, 1) if cap else 0.0

        return combo_summary

    @staticmethod
    def _course_code(course_name: str) -> str:
        parts = re.findall(r"[A-Za-z0-9]+", str(course_name))
        if not parts:
            return "CRS"
        if len(parts) == 1:
            return parts[0][:3].upper()
        return "".join(p[0].upper() for p in parts if p)

    def _combo_section_label(self, student: dict) -> str:
        """
        Build section label from G1 + G2 + base section, e.g. AI-IOT-A.
        If G2 is missing (single mode/partial), label falls back to G1-A.
        """
        g1 = student.get("allocated_primary")
        g2 = student.get("allocated_secondary")
        section = student.get("allocated_section", "")

        g1_code = self._course_code(g1) if g1 else "UNK"
        if g2:
            g2_code = self._course_code(g2)
            return f"{g1_code}-{g2_code}-{section}"
        return f"{g1_code}-{section}"

    @staticmethod
    def _safe_sheet_name(name: str) -> str:
        """Excel sheet names are max 31 chars and disallow []:*?/\\."""
        safe = re.sub(r"[\[\]:*?/\\]", "-", str(name)).strip()
        return safe[:31] if len(safe) > 31 else safe

    # ── Styling helpers ───────────────────────────────────────────────────────

    def _write_header_row(self, ws, headers: list[str]) -> None:
        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _CENTER
            cell.border = _BORDER
        ws.row_dimensions[1].height = 22

    def _style_data_row(
        self,
        ws,
        row_num: int,
        col_count: int,
        fill: PatternFill | None,
    ) -> None:
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = _BORDER
            cell.alignment = _LEFT if col_idx in (2, 4, 8) else _CENTER
            if fill:
                cell.fill = fill

    def _set_column_widths(self, ws, widths: list[int]) -> None:
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
